#!/usr/bin/env python3
"""
Trading Desk — Financial Records Import Pipeline

Stages:
  1. INGEST   → Files dropped into staging/inbox/
  2. VALIDATE → Check format, required columns, parseability
  3. TRANSFORM→ Convert rows to Beancount draft entries
  4. REVIEW   → Move to staging/pending/ for manual inspection or auto-approve
  5. IMPORT   → Append approved drafts to ledger/main.beancount
  6. ARCHIVE  → Move source file to staging/approved/ or staging/rejected/

Rollback:
  - Pre-import ledger backup is always created in ledger/.backups/
  - Import jobs are tracked in logs/import_jobs.jsonl
  - Rollback: restore ledger from backup + remove job from ledger

Supported formats: CSV, LibreOffice Calc (ods), Microsoft Excel (xlsx), PDF (Schwab statements)
"""

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import sys
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, List, Optional

try:
    from openpyxl import load_workbook
    openpyxl = True
except ImportError:
    openpyxl = None

try:
    import pdfplumber
    pdfplumber = True
except ImportError:
    pdfplumber = None

STAGING_ROOT = Path(__file__).resolve().parent.parent / "staging"
LEDGER_PATH = Path(__file__).resolve().parent.parent / "ledger" / "main.beancount"
BACKUPS_DIR = Path(__file__).resolve().parent.parent / "ledger" / ".backups"
LOGS_DIR = Path(__file__).resolve().parent.parent / "logs"
JOB_LOG = LOGS_DIR / "import_jobs.jsonl"
PIPE_LOG = LOGS_DIR / "pipeline.log"

US_EASTERN = "America/New_York"


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class ImportJob:
    job_id: str
    source_file: str
    status: str  # 'pending', 'approved', 'imported', 'rejected', 'rolled_back'
    started_at: str
    validated_at: Optional[str] = None
    imported_at: Optional[str] = None
    backed_up_ledger: Optional[str] = None
    draft_entries: List[str] = field(default_factory=list)
    ledger_start_line: Optional[int] = None
    ledger_end_line: Optional[int] = None
    error_message: Optional[str] = None


@dataclass
class ParsedRow:
    date: str
    description: str
    account: str
    amount: str
    commodity: str
    cost_spec: Optional[str] = None
    trade_id: Optional[str] = None


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def log_event(msg: str) -> None:
    ts = now_iso()
    line = f"[{ts}] {msg}\n"
    PIPE_LOG.parent.mkdir(parents=True, exist_ok=True)
    with PIPE_LOG.open("a", encoding="utf-8") as f:
        f.write(line)
    print(line.strip())


def generate_job_id(filepath: Path) -> str:
    h = hashlib.sha256(f"{filepath.name}{now_iso()}".encode()).hexdigest()[:12]
    return f"IMP-{h}"


def ensure_dirs() -> None:
    for d in ["inbox", "pending", "approved", "rejected"]:
        (STAGING_ROOT / d).mkdir(parents=True, exist_ok=True)
    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
    LOGS_DIR.mkdir(parents=True, exist_ok=True)


def load_jobs() -> List[dict]:
    if not JOB_LOG.exists():
        return []
    entries = []
    with JOB_LOG.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                entries.append(json.loads(line))
    return entries


def append_job(job: ImportJob) -> None:
    with JOB_LOG.open("a", encoding="utf-8") as f:
        f.write(json.dumps(asdict(job), default=str) + "\n")


def update_job_field(job_id: str, **kwargs) -> None:
    jobs = load_jobs()
    updated = []
    found = False
    for j in jobs:
        if j.get("job_id") == job_id:
            for k, v in kwargs.items():
                j[k] = v
            found = True
        updated.append(j)
    if not found:
        raise ValueError(f"Job {job_id} not found")
    with JOB_LOG.open("w", encoding="utf-8") as f:
        for j in updated:
            f.write(json.dumps(j, default=str) + "\n")


def line_count(path: Path) -> int:
    with path.open("r", encoding="utf-8") as f:
        return sum(1 for _ in f)


# ---------------------------------------------------------------------------
# Ingest
# ---------------------------------------------------------------------------

def ingest_file(src: Path) -> Path:
    ensure_dirs()
    dst = STAGING_ROOT / "inbox" / src.name
    if dst.exists():
        dst = dst.with_name(f"{src.stem}_{now_iso().replace(':','')}{src.suffix}")
    shutil.copy2(src, dst)
    log_event(f"INGEST: {src} -> {dst}")
    return dst


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

REQUIRED_CSV_HEADERS = {"date", "description", "account", "amount", "commodity"}


def sniff_delimiter(sample: str) -> str:
    for delim in [",", "\t", ";", "|"]:
        if delim in sample:
            return delim
    return ","


def validate_csv(path: Path) -> List[ParsedRow]:
    with path.open("r", encoding="utf-8", errors="replace") as f:
        sample = f.read(4096)
    delimiter = sniff_delimiter(sample)
    with path.open("r", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        headers = set(reader.fieldnames)
        if not REQUIRED_CSV_HEADERS.issubset(headers):
            missing = REQUIRED_CSV_HEADERS - headers
            raise ValueError(f"Missing required headers: {missing}")
    rows = []
    with path.open("r", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        for i, row in enumerate(reader, start=2):
            try:
                rows.append(
                    ParsedRow(
                        date=row["date"].strip(),
                        description=row["description"].strip(),
                        account=row["account"].strip(),
                        amount=row["amount"].strip(),
                        commodity=row["commodity"].strip(),
                        cost_spec=row.get("cost_spec", "").strip() or None,
                        trade_id=row.get("trade_id", "").strip() or None,
                    )
                )
            except Exception as e:
                raise ValueError(f"Row {i}: {e}")
    return rows


def validate_xlsx(path: Path) -> List[ParsedRow]:
    if openpyxl is None:
        raise ImportError("openpyxl is required for .xlsx files. Install: pip install openpyxl")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    if not REQUIRED_CSV_HEADERS.issubset(set(headers)):
        missing = REQUIRED_CSV_HEADERS - set(headers)
        raise ValueError(f"Missing required headers: {missing}")
    col_index = {h: i for i, h in enumerate(headers)}
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            rows.append(
                ParsedRow(
                    date=str(row[col_index["date"]]).strip(),
                    description=str(row[col_index["description"]]).strip(),
                    account=str(row[col_index["account"]]).strip(),
                    amount=str(row[col_index["amount"]]).strip(),
                    commodity=str(row[col_index["commodity"]]).strip(),
                    cost_spec=str(row[col_index["cost_spec"]]).strip() if "cost_spec" in col_index and row[col_index["cost_spec"]] else None,
                    trade_id=str(row[col_index["trade_id"]]).strip() if "trade_id" in col_index and row[col_index["trade_id"]] else None,
                )
            )
        except Exception as e:
            raise ValueError(f"Row {i}: {e}")
    return rows


def validate_file(path: Path) -> List[ParsedRow]:
    if not path.exists():
        raise FileNotFoundError(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return validate_csv(path)
    elif suffix in (".xlsx", ".ods"):
        return validate_xlsx(path)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")


# ---------------------------------------------------------------------------
# PDF parsing (Schwab statements)
# ---------------------------------------------------------------------------

def validate_pdf(path: Path, patches: Optional[dict] = None) -> List[str]:
    """Parse a Charles Schwab brokerage-statement PDF into Beancount draft entries."""
    if pdfplumber is None:
        raise ImportError(
            "pdfplumber is required for PDF parsing. "
            "Install: .venv/bin/pip install pdfplumber"
        )

    COLS = {
        "date": (0, 35),
        "category": (35, 75),
        "action": (75, 150),
        "description": (150, 430),
        "qty": (430, 500),
        "price": (500, 565),
        "fees": (565, 630),
        "amount": (630, 720),
        "gain": (720, 999),
    }
    CATEGORIES = {"Sale", "Purchase", "Deposit", "Withdrawal", "Dividend", "Interest", "Transfer"}
    ACTIONS = {"ShortSale", "CoverShort", "MiscCashCredit", "MiscCashDebit"}

    def group_words_by_line(words, tol=3):
        words = sorted(words, key=lambda w: float(w["top"]))
        lines = []
        current = []
        for w in words:
            if not current:
                current.append(w)
            else:
                if abs(float(w["top"]) - float(current[-1]["top"])) <= tol:
                    current.append(w)
                else:
                    lines.append(current)
                    current = [w]
        if current:
            lines.append(current)
        return lines

    def _num(s: str) -> float:
        s = s.replace(",", "").replace("$", "").strip()
        negative = s.startswith("(") and s.endswith(")")
        s = s.replace("(", "").replace(")", "")
        val = float(s) if s else 0.0
        return -val if negative else val

    def instrument_symbol(base: str, description: str) -> str:
        desc_upper = description.upper()
        opt_type = None
        if "CALL" in desc_upper:
            opt_type = "C"
        elif "PUT" in desc_upper:
            opt_type = "P"
        if not opt_type:
            return base
        strike_match = re.search(r"[$](\d+(?:\.\d+)?)", desc_upper)
        strike = int(float(strike_match.group(1))) if strike_match else None
        date_str = None
        m1 = re.search(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", desc_upper)
        if m1:
            mm, dd, yy = m1.groups()
            yy = yy[-2:]
            date_str = f"{yy}{mm.zfill(2)}{dd.zfill(2)}"
        else:
            m2 = re.search(r"(\d{1,2})\s+([A-Z]{3})\s+(\d{2,4})", desc_upper)
            if m2:
                dd, mon, yy = m2.groups()
                mon_map = {k: f"{(i+1):02d}" for i, k in enumerate(
                    ["JAN", "FEB", "MAR", "APR", "MAY", "JUN",
                     "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
                )}
                mm = mon_map.get(mon, "00")
                dd = dd.zfill(2)
                yy = yy[-2:]
                date_str = f"{yy}{mm}{dd}"
        if date_str and strike:
            return f"{base}{date_str}{opt_type}{strike}"
        elif strike:
            return f"{base}{opt_type}{strike}"
        return base

    transactions = []
    current_date = None
    year = datetime.now(timezone.utc).year
    m = re.search(r'(\d{4})', path.name)
    if m:
        year = int(m.group(1))

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if "Transaction Details" not in page_text:
                continue
            words = page.extract_words()
            line_groups = group_words_by_line(words)
            for line in line_groups:
                line.sort(key=lambda w: float(w["x0"]))
                fields = defaultdict(list)
                for w in line:
                    x = float(w["x0"])
                    for col, (lo, hi) in COLS.items():
                        if lo <= x < hi:
                            fields[col].append(w["text"])
                            break

                cat_candidates = [t for t in fields["category"] if t in CATEGORIES]
                if not cat_candidates:
                    continue
                category = cat_candidates[0]
                action_candidates = [t for t in fields["action"] if t in ACTIONS]
                action = action_candidates[0] if action_candidates else None
                date_word = fields["date"][0] if fields["date"] else None
                if date_word:
                    current_date = date_word

                amount_str = " ".join(fields["amount"])
                if not amount_str or amount_str.lower().startswith("total"):
                    continue

                transactions.append({
                    "date": current_date,
                    "category": category,
                    "action": action,
                    "description": " ".join(fields["description"]),
                    "qty": " ".join(fields["qty"]),
                    "price": " ".join(fields["price"]),
                    "fees": " ".join(fields["fees"]),
                    "amount": amount_str,
                    "gain": " ".join(fields["gain"]),
                })

    entries = []
    entries.append(f"\n;; --- Import batch: {path.name} ---")
    entries.append(f";; Import timestamp: {now_iso()}")
    entries.append(";; Source: Schwab brokerage statement")

    seq = 0
    last_date = None

    # In-memory FIFO lot trackers
    long_inventory: dict[str, list[tuple[float, float]]] = {}
    short_inventory: dict[str, list[tuple[float, float]]] = {}

    def _add(inv, symbol, qty, cost_per):
        if symbol not in inv:
            inv[symbol] = []
        inv[symbol].append((qty, cost_per))

    def _reduce(inv, symbol, qty):
        """Reduce qty FIFO. Returns list of (consumed_qty, cost_per)."""
        consumed = []
        remaining = qty
        lots = inv.get(symbol, [])
        while remaining > 0 and lots:
            lot_qty, lot_cost = lots[0]
            if lot_qty <= remaining:
                consumed.append((lot_qty, lot_cost))
                remaining -= lot_qty
                lots.pop(0)
            else:
                consumed.append((remaining, lot_cost))
                lots[0] = (lot_qty - remaining, lot_cost)
                remaining = 0
        inv[symbol] = lots
        return consumed, remaining

    # Pre-load existing ledger positions into FIFO trackers so sales/covers
    # can match lots established in prior statement periods.
    try:
        from beancount import loader as _loader
        _bc_entries, _, _ = _loader.load_file(str(LEDGER_PATH))
        for _entry in _bc_entries:
            if isinstance(_entry, _loader.data.Transaction):
                for posting in _entry.postings:
                    if posting.cost and posting.units.number != 0:
                        qty = float(posting.units.number)
                        cost = float(posting.cost.number)
                        symbol = str(posting.units.currency)
                        if posting.account == "Assets:Brokerage:Positions":
                            if qty < 0:
                                _reduce(long_inventory, symbol, abs(qty))
                            else:
                                _add(long_inventory, symbol, qty, cost)
                        elif posting.account == "Assets:Brokerage:ShortPositions":
                            if qty > 0:
                                _reduce(short_inventory, symbol, qty)
                            else:
                                _add(short_inventory, symbol, abs(qty), cost)
    except Exception:
        pass

    for t in transactions:
        date_str = t["date"]
        if date_str:
            mm, dd = date_str.split("/")
            date = f"{year}-{mm}-{dd}"
            last_date = date
        else:
            date = last_date
        if not date:
            continue

        seq += 1
        trade_id = f"SCHWAB-{date.replace('-', '')}-{seq:03d}"
        desc = f"{t['action'] or t['category']} {t['description']}".strip()

        qty = _num(t["qty"])
        price = _num(t["price"])
        fees = _num(t["fees"])
        amount = _num(t["amount"])
        gain_raw = re.sub(r"\(ST\)|\(LT\)|,", "", t["gain"]).strip()
        # Ignore Schwab-reported gain for lot-level transactions;
        # we derive realized P&L from our own FIFO inventory.

        base_symbol_match = re.search(r"^([A-Z]{1,5})", t["description"].upper())
        base_symbol = base_symbol_match.group(1) if base_symbol_match else "UNKNOWN"
        symbol = instrument_symbol(base_symbol, t["description"])

        lines = [f'{date} * "Trade #{trade_id}: {desc}"']
        cat = t["category"]
        act = t["action"]

        if cat in ("Deposit", "Withdrawal"):
            lines.append(f"  Assets:Brokerage:Cash    {amount:.2f} USD")
            lines.append(f"  Equity:Trading-Capital    {-amount:.2f} USD")
            entries.append("\n".join(lines))
            continue

        # Determine position account and flow direction
        if cat == "Purchase" and act != "CoverShort":
            pos_account = "Assets:Brokerage:Positions"
            gross = abs(amount) - fees
            unit_cost = gross / abs(qty) if qty else 0.0
            lines.append(f"  {pos_account}    {abs(qty):.0f} {symbol} {{{unit_cost:.4f} USD}}")
            lines.append(f"  Assets:Brokerage:Cash    {amount:.2f} USD")
            lines.append(f"  Expenses:Trading:Commissions    {fees:.2f} USD")
            _add(long_inventory, symbol, abs(qty), unit_cost)

        elif cat == "Sale" and act != "ShortSale":
            pos_account = "Assets:Brokerage:Positions"
            consumed, leftover = _reduce(long_inventory, symbol, abs(qty))
            total_basis = 0.0
            computed_cost = 0.0
            if leftover > 0:
                log_event(f"SKIP {trade_id}: No prior long lots for {symbol} (qty={abs(qty)}). Manual review required.")
                continue
            for c_qty, c_cost in consumed:
                lines.append(f"  {pos_account}    {-c_qty:.0f} {symbol} {{{c_cost:.4f} USD}}")
                total_basis += c_qty * c_cost
            cash_in = abs(amount)
            lines.append(f"  Assets:Brokerage:Cash    {cash_in:.2f} USD")
            lines.append(f"  Expenses:Trading:Commissions    {fees:.2f} USD")
            plug = -(-total_basis + cash_in + fees)
            if abs(plug) >= 0.01:
                lines.append(f"  Income:Trading:Realized-Gains    {plug:.2f} USD")

        elif cat == "ShortSale" or (act == "ShortSale"):
            pos_account = "Assets:Brokerage:ShortPositions"
            gross = abs(amount) + fees
            unit_cost = gross / abs(qty) if qty else 0.0
            lines.append(f"  {pos_account}    {qty:.0f} {symbol} {{{unit_cost:.4f} USD}}")
            lines.append(f"  Assets:Brokerage:Cash    {amount:.2f} USD")
            lines.append(f"  Expenses:Trading:Commissions    {fees:.2f} USD")
            _add(short_inventory, symbol, abs(qty), unit_cost)

        elif cat == "Purchase" and act == "CoverShort":
            pos_account = "Assets:Brokerage:ShortPositions"
            consumed, leftover = _reduce(short_inventory, symbol, abs(qty))
            total_basis = 0.0
            computed_cost = 0.0
            if leftover > 0:
                log_event(f"SKIP {trade_id}: No prior short lots for {symbol} (qty={abs(qty)}). Manual review required.")
                continue
            for c_qty, c_cost in consumed:
                lines.append(f"  {pos_account}    {c_qty:.0f} {symbol} {{{c_cost:.4f} USD}}")
                total_basis += c_qty * c_cost
            cash_out = amount
            lines.append(f"  Assets:Brokerage:Cash    {cash_out:.2f} USD")
            lines.append(f"  Expenses:Trading:Commissions    {fees:.2f} USD")
            plug = -(total_basis + cash_out + fees)
            if abs(plug) >= 0.01:
                lines.append(f"  Income:Trading:Realized-Gains    {plug:.2f} USD")

        elif cat in ("Interest",):
            lines.append(f"  Assets:Brokerage:Cash    {amount:.2f} USD")
            lines.append(f"  Income:Trading:Interest    {-amount:.2f} USD")

        elif cat in ("Dividend",):
            lines.append(f"  Assets:Brokerage:Cash    {amount:.2f} USD")
            lines.append(f"  Income:Trading:Dividends    {-amount:.2f} USD")

        else:
            lines.append(f"  ;; Unmapped {cat} / {act}: {desc}")
            lines.append(f"  Assets:Brokerage:Cash    {amount:.2f} USD")

        # Apply per-record patches (e.g. manual corrections for edge cases)
        if patches and trade_id in patches:
            rec_patch = patches[trade_id]
            if rec_patch.get("skip"):
                log_event(f"PATCH {trade_id}: record skipped per patch file")
                continue
            if rec_patch.get("postings"):
                # Rebuild entry with custom postings while keeping the header
                custom = [lines[0]]
                custom.extend(rec_patch["postings"])
                lines = custom
                log_event(f"PATCH {trade_id}: postings overridden")
            if rec_patch.get("replace_narration"):
                old_narr = lines[0]
                new_narr = old_narr.split('"')[0] + '"' + rec_patch["replace_narration"] + '"'
                lines[0] = new_narr
                log_event(f"PATCH {trade_id}: narration replaced")

        entries.append("\n".join(lines))

    return entries


# ---------------------------------------------------------------------------
# Transform
# ---------------------------------------------------------------------------

def transform_to_entries(rows: List[ParsedRow], source_file: str) -> List[str]:
    entries = []
    entries.append(f"\n;; --- Import batch: {source_file} ---")
    entries.append(f";; Import timestamp: {now_iso()}")

    # Group rows by trade_id (or date+description if no trade_id) into single transactions
    from itertools import groupby
    from operator import attrgetter

    def key_fn(r: ParsedRow) -> tuple:
        return (r.date, r.trade_id or r.description)

    rows = sorted(rows, key=key_fn)
    for (date, group_key), group in groupby(rows, key=key_fn):
        group_rows = list(group)
        # Use the first row's description/trade_id for the transaction header
        first = group_rows[0]
        if first.trade_id:
            desc = f'"Trade #{first.trade_id}: {first.description}"'
        else:
            desc = f'"{first.description}"'
        lines = [f"{date} * {desc}"]
        for r in group_rows:
            cost = f" {{{r.cost_spec}}}" if r.cost_spec else ""
            lines.append(f"  {r.account}    {r.amount} {r.commodity}{cost}")
        entries.append("\n".join(lines))
    return entries


# ---------------------------------------------------------------------------
# Ledger operations
# ---------------------------------------------------------------------------

def backup_ledger() -> Path:
    ensure_dirs()
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    backup = BACKUPS_DIR / f"main_{ts}.beancount"
    shutil.copy2(LEDGER_PATH, backup)
    log_event(f"BACKUP: {backup}")
    return backup


def append_to_ledger(entries: List[str]) -> tuple[int, int]:
    start_line = line_count(LEDGER_PATH) + 1
    with LEDGER_PATH.open("a", encoding="utf-8") as f:
        for e in entries:
            f.write(e + "\n")
    end_line = line_count(LEDGER_PATH)
    return start_line, end_line


def rollback_job(job_id: str) -> None:
    jobs = load_jobs()
    job = next((j for j in jobs if j.get("job_id") == job_id), None)
    if not job:
        raise ValueError(f"Job {job_id} not found")
    if job.get("status") != "imported":
        raise RuntimeError(f"Job {job_id} is not imported (status={job.get('status')})")

    backup_path = Path(job.get("backed_up_ledger", ""))
    start_line = job.get("ledger_start_line")
    end_line = job.get("ledger_end_line")

    if backup_path.exists():
        shutil.copy2(backup_path, LEDGER_PATH)
        log_event(f"ROLLBACK: Restored ledger from {backup_path}")
    elif start_line and end_line:
        lines = []
        with LEDGER_PATH.open("r", encoding="utf-8") as f:
            lines = f.readlines()
        new_lines = lines[: start_line - 1] + lines[end_line:]
        with LEDGER_PATH.open("w", encoding="utf-8") as f:
            f.writelines(new_lines)
        log_event(f"ROLLBACK: Removed lines {start_line}-{end_line}")
    else:
        raise RuntimeError("No backup or line range available for rollback")

    update_job_field(job_id, status="rolled_back", rolled_back_at=now_iso())
    log_event(f"ROLLBACK complete for {job_id}")


# ---------------------------------------------------------------------------
# Pipeline stages
# ---------------------------------------------------------------------------

def stage_validate(job: ImportJob, move_to_pending: bool = True, patches: Optional[dict] = None) -> ImportJob:
    src = Path(job.source_file)
    try:
        suffix = src.suffix.lower()
        if suffix == ".pdf":
            job.draft_entries = validate_pdf(src, patches=patches)
            job.status = "pending"
            job.validated_at = now_iso()
            log_event(f"VALIDATE OK: {job.job_id} ({len(job.draft_entries)} draft entries)")
        else:
            rows = validate_file(src)
            job.status = "pending"
            job.validated_at = now_iso()
            job.draft_entries = transform_to_entries(rows, src.name)
            log_event(f"VALIDATE OK: {job.job_id} ({len(rows)} rows)")
        if move_to_pending:
            dst = STAGING_ROOT / "pending" / src.name
            shutil.move(str(src), str(dst))
            job.source_file = str(dst)
            log_event(f"MOVED -> pending: {dst}")
    except Exception as e:
        job.status = "rejected"
        job.error_message = str(e)
        log_event(f"VALIDATE FAIL: {job.job_id} — {e}")
        dst = STAGING_ROOT / "rejected" / src.name
        shutil.move(str(src), str(dst))
        job.source_file = str(dst)
    append_job(job)
    return job


def stage_import(job_id: str, auto_approve: bool = False) -> None:
    jobs = load_jobs()
    job = next((j for j in jobs if j.get("job_id") == job_id), None)
    if not job:
        raise ValueError(f"Job {job_id} not found")
    if job.get("status") != "pending" and not auto_approve:
        raise RuntimeError(f"Job {job_id} is not pending (status={job.get('status')})")

    backup = backup_ledger()
    start_line, end_line = append_to_ledger(job["draft_entries"])
    job["status"] = "imported"
    job["imported_at"] = now_iso()
    job["backed_up_ledger"] = str(backup)
    job["ledger_start_line"] = start_line
    job["ledger_end_line"] = end_line
    update_job_field(**job)
    # Move source to approved
    src = Path(job["source_file"])
    dst = STAGING_ROOT / "approved" / src.name
    shutil.move(str(src), str(dst))
    job["source_file"] = str(dst)
    update_job_field(job_id, source_file=str(dst))
    log_event(f"IMPORT OK: {job_id} -> lines {start_line}-{end_line}")


def run_pipeline(src: Path, auto_approve: bool = False, patches: Optional[dict] = None) -> str:
    ensure_dirs()
    dst = ingest_file(src)
    job = ImportJob(
        job_id=generate_job_id(dst),
        source_file=str(dst),
        status="ingested",
        started_at=now_iso(),
    )
    log_event(f"JOB START: {job.job_id}")
    job = stage_validate(job, move_to_pending=True, patches=patches)
    if job.status == "pending" and auto_approve:
        stage_import(job.job_id, auto_approve=True)
    return job.job_id


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(description="Trading Desk Financial Records Import Pipeline")
    sub = parser.add_subparsers(dest="cmd")

    p_ingest = sub.add_parser("ingest", help="Ingest + validate a file")
    p_ingest.add_argument("file", type=Path, help="Path to CSV/XLSX/ODS/PDF")
    p_ingest.add_argument("--auto-approve", action="store_true", help="Auto-import after validation")
    p_ingest.add_argument("--patch-file", type=Path, required=True, help="JSON file mapping trade_ids to per-record patches")

    p_import = sub.add_parser("import", help="Import a pending job")
    p_import.add_argument("job_id", help="Job ID to import")

    p_rollback = sub.add_parser("rollback", help="Rollback an imported job")
    p_rollback.add_argument("job_id", help="Job ID to roll back")

    p_status = sub.add_parser("status", help="Show job statuses")

    p_review = sub.add_parser("review", help="List pending jobs for review")

    args = parser.parse_args()

    if args.cmd == "ingest":
        patches = None
        if args.patch_file:
            with open(args.patch_file, "r", encoding="utf-8") as pf:
                patches = json.load(pf)
            log_event(f"PATCH FILE loaded: {args.patch_file} ({len(patches)} patches)")
        jid = run_pipeline(args.file, auto_approve=args.auto_approve, patches=patches)
        print(f"Job ID: {jid}")
        return 0

    if args.cmd == "import":
        stage_import(args.job_id)
        return 0

    if args.cmd == "rollback":
        rollback_job(args.job_id)
        return 0

    if args.cmd == "status":
        jobs = load_jobs()
        if not jobs:
            print("No jobs found.")
            return 0
        print(f"{'Job ID':<20} {'Status':<15} {'Source':<40} {'Import Time'}")
        for j in jobs:
            src = os.path.basename(j.get("source_file", "N/A"))
            print(f"{j['job_id']:<20} {j['status']:<15} {src:<40} {j.get('imported_at','')}")
        return 0

    if args.cmd == "review":
        jobs = [j for j in load_jobs() if j.get("status") == "pending"]
        if not jobs:
            print("No pending jobs.")
            return 0
        for j in jobs:
            print(f"\n{j['job_id']} — {os.path.basename(j['source_file'])}")
            print("-" * 60)
            for entry in j.get("draft_entries", [])[:10]:
                print(entry)
            if len(j.get("draft_entries", [])) > 10:
                print(f"... (+{len(j['draft_entries']) - 10} more entries)")
        return 0

    parser.print_help()
    return 0


if __name__ == "__main__":
    sys.exit(main())
